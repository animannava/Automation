import os
import time
from unittest import skipIf

import openpyxl

from openpyxl.cell import cell
from openpyxl.writer.excel import save_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.ie.webdriver import WebDriver
options = webdriver.ChromeOptions()
filename="C:\\Users\\anith\\Downloads\\download.xlsx"
driver=webdriver.Chrome()
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()
time.sleep(2)
driver.find_element(By.ID,"downloadButton").click()
#time.sleep(10)
driver.implicitly_wait(5)
book=openpyxl.load_workbook(filename)
sheet=book.active
for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        if i==j and j>1:
            sheet.cell(row=i,column=j).value="Anitha "
            print(sheet.cell(row=i,column=j).value)
save_workbook(book,filename)
inputfile=driver.find_element(By.ID, "fileinput")
inputfile.send_keys(filename)
time.sleep(5)
os.remove(filename)
